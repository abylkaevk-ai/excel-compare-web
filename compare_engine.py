# compare_engine.py
# ВЕБ-ВЕРСИЯ (без GUI)

import os
import re
import math
from functools import lru_cache

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


HIGH_MATCH_THRESHOLD = 0.92
MED_MATCH_THRESHOLD = 0.80
EXTRACT_CONF_HIGH = 0.72
EXTRACT_CONF_MED = 0.45

MAX_HEADER_SCAN_ROWS = 140
WINDOW_STEP_ROWS = 35
WINDOW_HEIGHT = 25
MAX_BODY_ROWS = 200000
BLANK_STREAK_STOP = 25
HARD_MAX_COLS = 400
BLOCK_IF_NUMBER_SIG_CONFLICT = True


UNIT_SET = {
    "м", "м2", "м²", "м3", "м³", "шт", "кг", "т", "л",
    "п.м", "пог.м", "компл", "упак", "пач", "рул",
    "смена", "час", "маш.час", "маш.-час",
    "м/п", "м.п", "маш.ч", "маш.час."
}


STOP_WORDS = {
    "устройство", "монтаж", "демонтаж", "установка",
    "работы", "работ", "поставка", "материал",
    "материалы", "изделие", "изделия",
    "комплект", "по", "для", "на"
}


def safe_basename(path: str) -> str:
    return os.path.basename(path)


def is_blank_cell(v):
    return v is None or str(v).strip() == ""


@lru_cache(maxsize=250000)
def _norm_text_str(s: str):
    s = s.strip().lower().replace("ё", "е")
    s = re.sub(r"[^0-9a-zа-я\s\-\./×x]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def norm_text(x):
    if x is None:
        return ""
    return _norm_text_str(str(x))


def to_num(x):
    if x is None:
        return None
    try:
        return float(str(x).replace(",", "."))
    except:
        return None


def extract_numbers_signature(name):
    if not name:
        return tuple()
    nums = re.findall(r"\d+(?:\.\d+)?", norm_text(name))
    return tuple(sorted(set(nums)))


def smart_key(name):
    t = norm_text(name)
    tokens = [w for w in t.split() if w not in STOP_WORDS]
    sig = extract_numbers_signature(name)
    return " ".join(tokens) + "|" + ",".join(sig)


def extract_from_excel(path):
    wb = load_workbook(path, data_only=True)
    rows = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for r in ws.iter_rows(values_only=True):
            values = list(r)

            if not values or is_blank_cell(values[0]):
                continue

            name = str(values[0]).strip()
            qty = to_num(values[1]) if len(values) > 1 else None
            price = to_num(values[2]) if len(values) > 2 else None
            amount = to_num(values[3]) if len(values) > 3 else None

            rows.append({
                "file": safe_basename(path),
                "sheet": sheet_name,
                "name": name,
                "qty": qty,
                "price": price,
                "amount": amount,
                "smart_key": smart_key(name),
            })

    return rows, [], []


def collapse_duplicates(rows):
    grouped = {}
    for r in rows:
        key = (r["file"], r["smart_key"])
        if key not in grouped:
            grouped[key] = r
        else:
            if r["qty"] and grouped[key]["qty"]:
                grouped[key]["qty"] += r["qty"]
    return list(grouped.values())


def match_across_files(rows, files_order):
    by_key = {}

    for r in rows:
        by_key.setdefault(r["smart_key"], []).append(r)

    high = []
    unmatched = {f: [] for f in files_order}

    for k, items in by_key.items():
        files = {i["file"] for i in items}
        if len(files) >= 2:
            high.extend(items)
        else:
            unmatched[items[0]["file"]].extend(items)

    return high, [], unmatched


def build_excel_report(rows, review, diag, files_order, out_path):

    high, _, unmatched = match_across_files(rows, files_order)

    wb = Workbook()
    ws = wb.active
    ws.title = "Итог"

    ws.append(["Наименование", "Файл", "Кол-во", "Цена", "Сумма"])

    for r in high:
        ws.append([
            r["name"],
            r["file"],
            r["qty"],
            r["price"],
            r["amount"]
        ])

    wb.save(out_path)


def build_report(file_paths, out_path):

    files_order = [safe_basename(p) for p in file_paths]
    all_rows = []

    for p in file_paths:
        extracted, _, _ = extract_from_excel(p)
        extracted = collapse_duplicates(extracted)
        all_rows.extend(extracted)

    build_excel_report(all_rows, [], [], files_order, out_path)
